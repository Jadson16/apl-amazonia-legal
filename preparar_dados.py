"""
preparar_dados.py
Converte APLs_PA.xlsx → JSON de dados e simplifica GeoJSONs para uso web.
Execute uma vez sempre que a planilha for atualizada.
"""
import json
import unicodedata
from pathlib import Path

import geopandas as gpd
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd

# ── Configurações ──────────────────────────────────────────────────────────────
EXCEL_FILE = "APLs_PA.xlsx"
GEOJSON_DIR = Path("dados_ibge")
OUT_DATA = Path("data")
OUT_GEO = Path("geojson")
SIMPLIFY_TOLERANCE = 0.01   # graus; reduz tamanho do arquivo ~80%

OUT_DATA.mkdir(exist_ok=True)
OUT_GEO.mkdir(exist_ok=True)

# ── Mapeamento de regiões / rótulos ───────────────────────────────────────────
APL_LABELS = {
    "BaixoAmazonas":        "Baixo Amazonas",
    "Marajo":               "Marajó",
    "MetropolitanadeBelem": "Metropolitana de Belém",
    "NordesteParaense":     "Nordeste Paraense",
    "SudesteParaense":      "Sudeste Paraense",
    "SudoesteParaense":     "Sudoeste Paraense",
}

COR_FORA = "#c8c8c8"  # municípios fora de qualquer APL


def ylOrRd_hex(n: int) -> list[str]:
    """Retorna n cores hex da paleta YlOrRd do matplotlib, do mais claro ao mais escuro."""
    cmap = plt.cm.get_cmap("YlOrRd", n)
    cores = []
    for i in range(n):
        r, g, b, _ = cmap(i)
        cores.append("#{:02x}{:02x}{:02x}".format(int(r*255), int(g*255), int(b*255)))
    return cores


def norm(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn").strip().lower()


def apl_key(apl_raw: str) -> str:
    parts = apl_raw.split("-")
    middle = "-".join(parts[1:-1]) if len(parts) >= 3 else apl_raw
    return norm(middle).title().replace(" ", "")


def apl_label(apl_raw: str) -> str:
    key_norm = "".join(
        c for c in unicodedata.normalize("NFD", apl_key(apl_raw))
        if unicodedata.category(c) != "Mn"
    )
    for k, v in APL_LABELS.items():
        if norm(k) == norm(key_norm):
            return v
    return apl_key(apl_raw)


# ── Leitura da planilha ────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_FILE)
print(f"Abas encontradas: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    df = pd.DataFrame(rows, columns=["cadeia", "apl", "municipio", "idri", "count", "sum_val"])
    df = df[df["municipio"].notna()].copy()
    df["mun_norm"] = df["municipio"].apply(norm)
    df["apl_label"] = df["apl"].apply(apl_label)

    # IDRI médio por APL
    idri_medio = df.groupby("apl")["idri"].mean().to_dict()

    # Estado (2 últimas letras do nome do APL)
    estado = df["apl"].iloc[0].split("-")[-1] if not df.empty else "PA"
    cadeia = df["cadeia"].iloc[0] if not df.empty else sheet_name

    # Montar estrutura de APLs (sem cor ainda — atribuída após ordenar por IDRI)
    apls_out = []
    for apl_raw, grupo in df.groupby("apl"):
        label = apl_label(apl_raw)
        apls_out.append({
            "id": apl_raw,
            "label": label,
            "idri_medio": round(idri_medio[apl_raw], 5),
            "municipios": sorted(grupo["municipio"].unique().tolist()),
        })
    apls_out.sort(key=lambda x: x["idri_medio"])

    # Atribui cores YlOrRd na ordem crescente de IDRI (igual ao matplotlib)
    cores = ylOrRd_hex(len(apls_out))
    for i, apl in enumerate(apls_out):
        apl["cor"] = cores[i]

    # Montar dicionário por município
    muns_out = {}
    for _, row in df.iterrows():
        muns_out[row["municipio"]] = {
            "apl_id": row["apl"],
            "apl_label": row["apl_label"],
            "idri": round(row["idri"], 5),
        }

    resultado = {
        "estado": estado,
        "cadeia": cadeia,
        "apls": apls_out,
        "municipios": muns_out,
    }

    out_file = OUT_DATA / f"{norm(cadeia)}_{estado}.json"
    out_file.write_text(json.dumps(resultado, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  -> {out_file}  ({len(muns_out)} municipios, {len(apls_out)} APLs)")

# ── Simplificação dos GeoJSONs ────────────────────────────────────────────────
geo_sources = {
    "PA": GEOJSON_DIR / "PA_municipios_geo.geojson",
}

for estado, geo_path in geo_sources.items():
    if not geo_path.exists():
        print(f"  [aviso] GeoJSON não encontrado: {geo_path}")
        continue
    print(f"Simplificando GeoJSON {estado}...")
    gdf = gpd.read_file(geo_path)
    gdf_simple = gdf[["nome", "geometry"]].copy()
    gdf_simple["geometry"] = gdf_simple["geometry"].simplify(SIMPLIFY_TOLERANCE, preserve_topology=True)
    out_geo = OUT_GEO / f"{estado}.geojson"
    gdf_simple.to_file(out_geo, driver="GeoJSON")
    orig_kb = geo_path.stat().st_size // 1024
    simp_kb = out_geo.stat().st_size // 1024
    print(f"  -> {out_geo}  ({orig_kb} KB -> {simp_kb} KB)")

print("\nDados prontos!")
